#!/usr/bin/env python3
"""
Parser Worker - 文件解析器
支持多种文件格式：PDF, Word, Excel, TXT, Markdown, HTML 等
"""

import sys
import json
import logging
import chardet
from typing import Dict, Any, Optional
from pathlib import Path
from datetime import datetime

logger = logging.getLogger("parser_worker")


class FileParser:
    """文件解析器"""
    
    def __init__(self):
        self.supported_extensions = {
            '.txt': self._parse_txt,
            '.md': self._parse_txt,
            '.markdown': self._parse_txt,
            '.pdf': self._parse_pdf,
            '.doc': self._parse_doc,
            '.docx': self._parse_docx,
            '.xls': self._parse_xls,
            '.xlsx': self._parse_xlsx,
            '.html': self._parse_html,
            '.htm': self._parse_html,
            '.xml': self._parse_xml,
            '.json': self._parse_json,
            '.csv': self._parse_csv,
            '.py': self._parse_code,
            '.js': self._parse_code,
            '.ts': self._parse_code,
            '.java': self._parse_code,
            '.cpp': self._parse_code,
            '.c': self._parse_code,
            '.go': self._parse_code,
            '.rs': self._parse_code,
            '.rb': self._parse_code,
            '.php': self._parse_code,
            '.sh': self._parse_code,
            '.bash': self._parse_code,
        }
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        解析文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            {
                "content": str,  # 文件内容
                "metadata": {    # 元数据
                    "file_type": str,
                    "size": int,
                    "lines": int,
                    ...
                }
            }
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        ext = path.suffix.lower()
        
        if ext not in self.supported_extensions:
            # 尝试作为文本文件解析
            return self._parse_txt(file_path)
        
        parser_func = self.supported_extensions[ext]
        return parser_func(file_path)
    
    def _detect_encoding(self, file_path: str) -> str:
        """检测文件编码"""
        with open(file_path, 'rb') as f:
            raw = f.read(10000)  # 读取前 10KB
            result = chardet.detect(raw)
            return result['encoding'] or 'utf-8'
    
    def _parse_txt(self, file_path: str) -> Dict[str, Any]:
        """解析纯文本文件"""
        encoding = self._detect_encoding(file_path)
        
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()
        except UnicodeDecodeError:
            # 尝试其他编码
            with open(file_path, 'r', encoding='gbk', errors='ignore') as f:
                content = f.read()
        
        lines = content.split('\n')
        
        return {
            "content": content,
            "metadata": {
                "file_type": "text",
                "size": len(content),
                "lines": len(lines),
                "encoding": encoding
            }
        }
    
    def _parse_pdf(self, file_path: str) -> Dict[str, Any]:
        """解析 PDF 文件"""
        try:
            import fitz  # PyMuPDF
        except ImportError:
            raise ImportError("PyMuPDF not installed. Run: pip install PyMuPDF")
        
        doc = fitz.open(file_path)
        texts = []
        
        for page in doc:
            texts.append(page.get_text())
        
        doc.close()
        
        content = '\n'.join(texts)
        
        return {
            "content": content,
            "metadata": {
                "file_type": "pdf",
                "size": len(content),
                "pages": len(doc) if 'doc' in locals() else 0
            }
        }
    
    def _parse_docx(self, file_path: str) -> Dict[str, Any]:
        """解析 Word 文档"""
        try:
            from docx import Document
        except ImportError:
            raise ImportError("python-docx not installed. Run: pip install python-docx")
        
        doc = Document(file_path)
        texts = [para.text for para in doc.paragraphs if para.text.strip()]
        
        content = '\n'.join(texts)
        
        return {
            "content": content,
            "metadata": {
                "file_type": "docx",
                "size": len(content),
                "paragraphs": len(texts)
            }
        }
    
    def _parse_doc(self, file_path: str) -> Dict[str, Any]:
        """解析旧版 Word 文档 (.doc)"""
        # .doc 格式需要 antiword 或类似工具
        # 暂时返回错误
        return {
            "content": "",
            "metadata": {
                "file_type": "doc",
                "error": ".doc format not supported, please convert to .docx"
            }
        }
    
    def _parse_xlsx(self, file_path: str) -> Dict[str, Any]:
        """解析 Excel 文件"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        wb = load_workbook(file_path, data_only=True)
        texts = []
        
        for sheet in wb.worksheets:
            texts.append(f"=== Sheet: {sheet.title} ===")
            for row in sheet.iter_rows(values_only=True):
                row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                texts.append(row_text)
        
        content = '\n'.join(texts)
        
        return {
            "content": content,
            "metadata": {
                "file_type": "xlsx",
                "size": len(content),
                "sheets": len(wb.worksheets)
            }
        }
    
    def _parse_xls(self, file_path: str) -> Dict[str, Any]:
        """解析旧版 Excel 文件 (.xls)"""
        return {
            "content": "",
            "metadata": {
                "file_type": "xls",
                "error": ".xls format not supported, please convert to .xlsx"
            }
        }
    
    def _parse_html(self, file_path: str) -> Dict[str, Any]:
        """解析 HTML 文件"""
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            raise ImportError("beautifulsoup4 not installed. Run: pip install beautifulsoup4")
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        soup = BeautifulSoup(content, 'lxml')
        
        # 提取文本
        text = soup.get_text(separator='\n', strip=True)
        
        return {
            "content": text,
            "metadata": {
                "file_type": "html",
                "size": len(text),
                "title": soup.title.string if soup.title else ""
            }
        }
    
    def _parse_xml(self, file_path: str) -> Dict[str, Any]:
        """解析 XML 文件"""
        return self._parse_txt(file_path)
    
    def _parse_json(self, file_path: str) -> Dict[str, Any]:
        """解析 JSON 文件"""
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        return {
            "content": content,
            "metadata": {
                "file_type": "json",
                "size": len(content)
            }
        }
    
    def _parse_csv(self, file_path: str) -> Dict[str, Any]:
        """解析 CSV 文件"""
        return self._parse_txt(file_path)
    
    def _parse_code(self, file_path: str) -> Dict[str, Any]:
        """解析代码文件"""
        return self._parse_txt(file_path)


def process_request(request: Dict[str, Any]) -> Dict[str, Any]:
    """处理请求"""
    action = request.get("action")
    
    if action == "parse":
        file_path = request.get("file_path")
        if not file_path:
            return {"error": "file_path is required"}
        
        parser = FileParser()
        result = parser.parse(file_path)
        return result
    
    elif action == "test":
        file_path = request.get("file_path", "/tmp/test.txt")
        parser = FileParser()
        result = parser.parse(file_path)
        return {
            "status": "success",
            "file": file_path,
            "content_preview": result["content"][:200],
            "metadata": result["metadata"]
        }
    
    else:
        return {"error": f"Unknown action: {action}"}


def main():
    """主函数"""
    if len(sys.argv) > 1 and sys.argv[1] == "--test":
        # 命令行测试模式
        file_path = sys.argv[2] if len(sys.argv) > 2 else "/tmp/test.txt"
        request = {"action": "test", "file_path": file_path}
        result = process_request(request)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return
    
    # 守护进程模式
    logger.info("Parser worker started")
    
    parser = FileParser()
    
    for line in sys.stdin:
        try:
            request = json.loads(line.strip())
            action = request.get("action")
            
            if action == "parse":
                file_path = request.get("file_path")
                if not file_path:
                    print(json.dumps({"error": "file_path is required"}))
                    continue
                
                result = parser.parse(file_path)
                print(json.dumps(result, ensure_ascii=False))
            
            elif action == "quit":
                logger.info("Shutting down")
                print(json.dumps({"status": "shutdown"}))
                break
            
            else:
                print(json.dumps({"error": f"Unknown action: {action}"}))
            
            sys.stdout.flush()
            
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON: {e}")
            print(json.dumps({"error": f"Invalid JSON: {e}"}))
            sys.stdout.flush()
        except Exception as e:
            logger.error(f"Error processing request: {e}")
            print(json.dumps({"error": str(e)}))
            sys.stdout.flush()


if __name__ == "__main__":
    main()
